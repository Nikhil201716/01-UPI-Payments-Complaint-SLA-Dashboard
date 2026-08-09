"""
build_excel_dashboard.py
-------------------------
Builds a self-contained, stakeholder-ready Excel dashboard
(reports/UPI_Complaint_SLA_Dashboard.xlsx) directly from the SQLite
database. Designed for people who do NOT have Power BI/Tableau installed -
just Excel - following the same "Excel dashboard with clickable
navigation" pattern used on the Uber Supply-Demand Gap Analysis project.

Sheets produced:
  1. Dashboard        - KPI cards + charts + navigation links (front page)
  2. By Category       - breach-rate table, colour-scaled, with bar chart
  3. By Channel         - breach-rate table with bar chart
  4. Monthly Trend      - volume + SLA compliance trend with line chart
  5. Raw Data (filterable) - full ticket-level table with AutoFilter
"""

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "database" / "complaints.db"
OUT_PATH = ROOT / "reports" / "UPI_Complaint_SLA_Dashboard.xlsx"
OUT_PATH.parent.mkdir(exist_ok=True)

NAVY = "1F3A5F"
ACCENT = "2E6F40"
RED = "C0392B"
LIGHT_GREY = "F2F2F2"
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=NAVY, bold=True, size=18)
SUB_FONT = Font(color="555555", italic=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ----------------------------------------------------------------------
# 1. Pull data
# ----------------------------------------------------------------------
conn = sqlite3.connect(DB_PATH)

kpi = pd.read_sql_query("""
    SELECT
        COUNT(*) AS total_tickets,
        SUM(CASE WHEN status='Closed' THEN 1 ELSE 0 END) AS closed_tickets,
        SUM(CASE WHEN status='Open' THEN 1 ELSE 0 END) AS open_tickets,
        ROUND(100.0*SUM(CASE WHEN sla_breached=0 THEN 1 ELSE 0 END)/COUNT(*),1) AS sla_compliance_pct,
        ROUND(AVG(resolution_hours),1) AS avg_resolution_hours,
        ROUND(SUM(transaction_amount_inr),0) AS total_disputed_value_inr
    FROM fact_complaints
""", conn).iloc[0]

by_category = pd.read_sql_query("""
    SELECT c.category_name, c.category_group, c.sla_target_hours,
           COUNT(*) AS ticket_count,
           ROUND(AVG(f.resolution_hours),1) AS avg_resolution_hours,
           ROUND(100.0*SUM(f.sla_breached)/COUNT(*),1) AS sla_breach_pct
    FROM fact_complaints f JOIN dim_category c ON f.category_id=c.category_id
    GROUP BY c.category_name, c.category_group, c.sla_target_hours
    ORDER BY sla_breach_pct DESC
""", conn)

by_channel = pd.read_sql_query("""
    SELECT ch.channel_name, COUNT(*) AS ticket_count,
           ROUND(100.0*SUM(f.sla_breached)/COUNT(*),1) AS sla_breach_pct
    FROM fact_complaints f JOIN dim_channel ch ON f.channel_id=ch.channel_id
    GROUP BY ch.channel_name ORDER BY sla_breach_pct DESC
""", conn)

monthly = pd.read_sql_query("""
    SELECT strftime('%Y-%m', opened_at) AS month, COUNT(*) AS ticket_count,
           ROUND(100.0*SUM(CASE WHEN sla_breached=0 THEN 1 ELSE 0 END)/COUNT(*),1) AS sla_compliance_pct
    FROM fact_complaints GROUP BY month ORDER BY month
""", conn)

raw = pd.read_sql_query("""
    SELECT f.ticket_id, f.customer_id, c.category_name, c.category_group,
           ch.channel_name, a.agent_name, a.team, f.transaction_amount_inr,
           f.opened_at, f.closed_at, f.status, f.resolution_hours,
           f.sla_target_hours, f.sla_breached, f.root_cause, f.is_incident_day
    FROM fact_complaints f
    JOIN dim_category c ON f.category_id=c.category_id
    JOIN dim_channel ch ON f.channel_id=ch.channel_id
    JOIN dim_agent a ON f.agent_id=a.agent_id
    ORDER BY f.opened_at
""", conn)
conn.close()

# ----------------------------------------------------------------------
# 2. Build workbook
# ----------------------------------------------------------------------
wb = Workbook()


def style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header_row(ws, start_row, len(df.columns))
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    for j, col in enumerate(df.columns):
        width = max(12, min(45, int(df[col].astype(str).str.len().max() if len(df) else 10) + 2, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(start_col + j)].width = width
    return start_row + len(df)


# --- Sheet 1: Dashboard (front page) -----------------------------------
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False

ws["B2"] = "UPI Digital Payments — Complaint & SLA Analytics Dashboard"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Synthetic support-ticket data | 12-month view | Generated from complaints.db"
ws["B3"].font = SUB_FONT

kpi_labels = [
    ("Total Tickets", f"{int(kpi.total_tickets):,}"),
    ("SLA Compliance %", f"{kpi.sla_compliance_pct}%"),
    ("Avg Resolution (hrs)", f"{kpi.avg_resolution_hours}"),
    ("Open Tickets", f"{int(kpi.open_tickets):,}"),
    ("Disputed Value (INR)", f"₹{kpi.total_disputed_value_inr:,.0f}"),
]
col = 2
for label, value in kpi_labels:
    c1 = ws.cell(row=5, column=col, value=label)
    c1.font = Font(bold=True, color="FFFFFF", size=10)
    c1.fill = PatternFill("solid", fgColor=ACCENT)
    c1.alignment = Alignment(horizontal="center")
    c2 = ws.cell(row=6, column=col, value=value)
    c2.font = Font(bold=True, color=NAVY, size=14)
    c2.alignment = Alignment(horizontal="center")
    c2.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    ws.column_dimensions[get_column_letter(col)].width = 20
    col += 1

ws["B9"] = "Navigate to:"
ws["B9"].font = Font(bold=True, color=NAVY)
nav_links = [("By Category", "'By Category'!A1"), ("By Channel", "'By Channel'!A1"),
             ("Monthly Trend", "'Monthly Trend'!A1"), ("Raw Data", "'Raw Data'!A1")]
r = 10
for text, target in nav_links:
    cell = ws.cell(row=r, column=2, value=f"→ {text}")
    cell.hyperlink = f"#{target}"
    cell.font = Font(color="1155CC", underline="single")
    r += 1

# Mini category chart embedded on the dashboard sheet
cat_start = 16
ws.cell(row=cat_start - 1, column=2, value="SLA Breach % by Category").font = Font(bold=True, color=NAVY, size=12)
write_df(ws, by_category[["category_name", "sla_breach_pct"]].rename(
    columns={"category_name": "Category", "sla_breach_pct": "SLA Breach %"}), start_row=cat_start, start_col=2)

chart = BarChart()
chart.title = "SLA Breach % by Category"
chart.y_axis.title = "Breach %"
chart.style = 10
data = Reference(ws, min_col=3, min_row=cat_start, max_row=cat_start + len(by_category))
cats = Reference(ws, min_col=2, min_row=cat_start + 1, max_row=cat_start + len(by_category))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width, chart.height = 22, 10
ws.add_chart(chart, "H16")

# --- Sheet 2: By Category ----------------------------------------------
ws2 = wb.create_sheet("By Category")
end_row = write_df(ws2, by_category.rename(columns={
    "category_name": "Category", "category_group": "Group",
    "sla_target_hours": "SLA Target (hrs)", "ticket_count": "Tickets",
    "avg_resolution_hours": "Avg Resolution (hrs)", "sla_breach_pct": "SLA Breach %"}))
rule = ColorScaleRule(start_type="min", start_color="63BE7B", end_type="max", end_color="F8696B")
ws2.conditional_formatting.add(f"F2:F{end_row}", rule)

chart2 = BarChart()
chart2.title = "SLA Breach % by Category"
data2 = Reference(ws2, min_col=6, min_row=1, max_row=end_row)
cats2 = Reference(ws2, min_col=1, min_row=2, max_row=end_row)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width, chart2.height = 24, 11
ws2.add_chart(chart2, "H2")

# --- Sheet 3: By Channel -------------------------------------------------
ws3 = wb.create_sheet("By Channel")
end_row3 = write_df(ws3, by_channel.rename(columns={
    "channel_name": "Channel", "ticket_count": "Tickets", "sla_breach_pct": "SLA Breach %"}))
chart3 = BarChart()
chart3.title = "SLA Breach % by Channel"
data3 = Reference(ws3, min_col=3, min_row=1, max_row=end_row3)
cats3 = Reference(ws3, min_col=1, min_row=2, max_row=end_row3)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width, chart3.height = 20, 10
ws3.add_chart(chart3, "F2")

# --- Sheet 4: Monthly Trend ---------------------------------------------
ws4 = wb.create_sheet("Monthly Trend")
end_row4 = write_df(ws4, monthly.rename(columns={
    "month": "Month", "ticket_count": "Tickets", "sla_compliance_pct": "SLA Compliance %"}))
chart4 = LineChart()
chart4.title = "Monthly Ticket Volume & SLA Compliance %"
data4 = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=end_row4)
cats4 = Reference(ws4, min_col=1, min_row=2, max_row=end_row4)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.width, chart4.height = 24, 11
ws4.add_chart(chart4, "F2")

# --- Sheet 5: Raw Data (filterable) -------------------------------------
ws5 = wb.create_sheet("Raw Data")
end_row5 = write_df(ws5, raw)
tab = Table(displayName="RawComplaints", ref=f"A1:{get_column_letter(len(raw.columns))}{end_row5}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws5.add_table(tab)
ws5.freeze_panes = "A2"

# ----------------------------------------------------------------------
# 3. Save
# ----------------------------------------------------------------------
wb.save(OUT_PATH)
print(f"Excel dashboard saved to: {OUT_PATH}")
print(f"Rows in Raw Data sheet: {len(raw):,}")
